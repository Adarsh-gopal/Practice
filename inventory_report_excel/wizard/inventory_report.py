# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.



import io
import locale
import base64
import textwrap
from copy import copy

from datetime import datetime
from openpyxl import Workbook
from odoo import models, fields, api,_
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from odoo.exceptions import UserError, ValidationError,Warning
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from pytz import timezone
import pdb

class InventoryReportExcel(models.TransientModel):
    _name = "inventory.report.excel"
    _description = "inventory.report.excel"

    date_start = fields.Datetime(string="Start Date", required=True)
    date_end = fields.Datetime(string="End Date", required=True, default=fields.Datetime.now)
    inv_report = fields.Binary('Inventory Report')
    file_name = fields.Char('File Name')
    inv_printed = fields.Boolean('inventory Printed')
    warehouse_id = fields.Many2many('stock.warehouse')
    location_id = fields.Many2many('stock.location')
    category_id = fields.Many2one('product.category',string='Category')
    product_ids = fields.Many2many('product.product',string='Products')
    company_id = fields.Many2one('res.company',string='Company',required=True,default=lambda self: self.env.user.company_id)
    filter_by = fields.Selection([('product','Product'),('category','Product Category')],string='Product /Category', required=True)
    location_type = fields.Selection([('warehouse','Warehouse')],string='warehouse', required=True)
    

    @api.constrains('date_start')
    def _code_constrains(self):
        if self.date_start > self.date_end:
            raise ValidationError(_("'Start Date' must be before 'End Date'"))

    # Get the product and category filters 
    def get_product(self):
        product_pool=self.env['product.product']
        if not self.filter_by:
            product_ids = product_pool.search([('type','!=','service')])
            return product_ids
        elif self.filter_by == 'product' and self.product_ids:
            return self.product_ids
        elif self.filter_by == 'category' and self.category_id:
            product_ids = product_pool.search([('categ_id','child_of',self.category_id.id),('type','!=','service')])
            return product_ids


    

    def generate_report(self):

    
        # #Create Workbook and Worksheet
        # now_utc = self.date_start(timezone('UTC'))
        # format = "%Y-%m-%d %H:%M:%S %Z%z"
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"
        new_start=self.date_start.astimezone(timezone('Asia/Kolkata'))
        new_end=self.date_end.astimezone(timezone('Asia/Kolkata'))
        report_heading = " Inventory Report from" + ' ' + datetime.strftime(new_start, '%d-%m-%Y') + ' '+ 'To' + ' '+ datetime.strftime( new_end, '%d-%m-%Y')
        
        # #Border
        thin = Side(border_style="thin", color="000000")
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=18)

        comapny = ws.cell(row=1, column=1, value=self.env.user.company_id.name + report_heading)
        comapny.alignment = Alignment(horizontal='center',vertical='center')
        comapny.border = Border(bottom=thin,top=thin)
        comapny.font = Font(size=10,name='Calibri')


        # Defining the Table Caolumn headings
        sl_no = ws.cell(row=3, column=1, value="S.No")
        sl_no.alignment = copy(comapny.alignment)


        product_id = ws.cell(row=3, column=2, value="Product")
        product_id.alignment = copy(comapny.alignment)

        product_cat = ws.cell(row=3, column=3, value="Product Category")
        product_cat.alignment = Alignment(horizontal='center',vertical='center')

        product_code= ws.cell(row=3, column=4, value="Internal Reference")
        product_code.alignment = Alignment(horizontal='center',vertical='center')

        # location_name= ws.cell(row=3, column=5, value="Location")
        # location_name.alignment = Alignment(horizontal='center',vertical='center')

        p_unit = ws.cell(row=3, column=5, value="Unit of Measure")
        p_unit.alignment = Alignment(horizontal='center',vertical='center')

        sys_opening = ws.cell(row=3, column=6, value="Opening Qty")
        sys_opening.alignment = Alignment(horizontal='center',vertical='center')
        
        sys_amt = ws.cell(row=3, column=7, value="Opening Value")
        sys_amt.alignment = Alignment(horizontal='center',vertical='center')

        purchase_qty = ws.cell(row=3, column=8, value="Purchase Shipment Qty")
        purchase_qty.alignment = Alignment(horizontal='center',vertical='center')

        purchase_amt = ws.cell(row=3, column=9, value="Purchase Shipment Value")
        purchase_amt.alignment = Alignment(horizontal='center',vertical='center')

        purchase_r_qty = ws.cell(row=3, column=10, value="Purchase Return Qty")
        purchase_r_qty.alignment = Alignment(horizontal='center',vertical='center')

        purchase_r_amt = ws.cell(row=3, column=11, value="Purchase Return Value")
        purchase_r_amt.alignment = Alignment(horizontal='center',vertical='center')

        sale_qty = ws.cell(row=3, column=12, value="Sale Qty")
        sale_qty.alignment = Alignment(horizontal='center',vertical='center')

        sale_amt = ws.cell(row=3, column=13, value="Sale Value")
        sale_amt.alignment = Alignment(horizontal='center',vertical='center')

        sale_re_qty = ws.cell(row=3, column=14, value="Sale Return Qty")
        sale_re_qty.alignment = Alignment(horizontal='center',vertical='center')

        sale_re_amt = ws.cell(row=3, column=15, value="Sale Return Value")
        sale_re_amt.alignment = Alignment(horizontal='center',vertical='center')

        mrp_qty = ws.cell(row=3, column=16, value="Manufactured Qty")
        mrp_qty.alignment = Alignment(horizontal='center',vertical='center')

        mrp_amt = ws.cell(row=3, column=17, value="Manufactured Value")
        mrp_amt.alignment = Alignment(horizontal='center',vertical='center')

        consumed_qty = ws.cell(row=3, column=18, value="Consumed Qty")
        consumed_qty.alignment = Alignment(horizontal='center',vertical='center')

        consumed_amt = ws.cell(row=3, column=19, value="Consumed Value")
        consumed_amt.alignment = Alignment(horizontal='center',vertical='center')
 
        postive_qty = ws.cell(row=3, column=20, value="Postive Adjustment Qty")
        postive_qty.alignment = Alignment(horizontal='center',vertical='center')

        postive_amt = ws.cell(row=3, column=21, value="Postive Adjustment Value")
        postive_amt.alignment = Alignment(horizontal='center',vertical='center')

        negative_qty = ws.cell(row=3, column=22, value="Negative Adjustment Qty")
        negative_qty.alignment = Alignment(horizontal='center',vertical='center')

        negative_amt = ws.cell(row=3, column=23, value="Negative Adjustment Value")
        negative_amt.alignment = Alignment(horizontal='center',vertical='center')
        
        t_shipment_inter_qty = ws.cell(row=3, column=24, value="Transfer Shipment Qty")
        t_shipment_inter_qty.alignment = Alignment(horizontal='center',vertical='center')

        t_shipment_inter_amt = ws.cell(row=3, column=25, value="Transfer Shipment Value")
        t_shipment_inter_amt.alignment = Alignment(horizontal='center',vertical='center')

        # t_shipment_intra_qty = ws.cell(row=3, column=24, value="Intra Transfer Shipment Qty")
        # t_shipment_intra_qty.alignment = Alignment(horizontal='center',vertical='center')

        # t_shipment_intra_amt = ws.cell(row=3, column=25, value="Intra Transfer Shipment Value")
        # t_shipment_intra_amt.alignment = Alignment(horizontal='center',vertical='center')
        

        t_receipt_inter_qty = ws.cell(row=3, column=26, value="Transfer Receipt Qty")
        t_receipt_inter_qty.alignment = Alignment(horizontal='center',vertical='center')

        t_receipt_inter_amt = ws.cell(row=3, column=27, value="Transfer Receipt Value")
        t_receipt_inter_amt.alignment = Alignment(horizontal='center',vertical='center')


        # t_receipt_intra_qty = ws.cell(row=3, column=28, value="Intra Transfer Receipt Qty")
        # t_receipt_intra_qty.alignment = Alignment(horizontal='center',vertical='center')

        # t_receipt_intra_amt = ws.cell(row=3, column=29, value="Intra Transfer Receipt Value")
        # t_receipt_intra_amt.alignment = Alignment(horizontal='center',vertical='center')


        closing_qty = ws.cell(row=3, column=28, value="Closing Stock Qty")
        closing_qty.alignment = Alignment(horizontal='center',vertical='center')

        closing_amt = ws.cell(row=3, column=29, value="Closing Stock Value")
        closing_amt.alignment = Alignment(horizontal='center',vertical='center')

        # Purchase  = Purchase Receipt - Purchase Return
        # Sales = Sales Shipment - Sales Return
        # Manufacture = Consumpution-Output
        # Consumed = Transfer Shipment- Transfer Receipt
        # Postive Adjustmen = 
        # Negative Adjustment
        
        
        # caliculating the opening qty.
        opening_qty_query = '''
            select 
                    product_id ,
                    ABS(sum(product_qty)) as qty,
                    ABS(sum(value)) as value
            from 
                    inventory_base_report 
            where 
                    date < %s and 
                    ( location_id  IN %s or 
                    warehouse_id IN %s ) and
                     product_id =%s  and 
                     transaction_types NOT IN  ( 'internal' )
            group by product_id


            '''

        # caliculating the opening value.
        opening_value_query = '''
            select 
                    product_id ,
                    ABS(sum(product_qty)) as qty,
                    ABS(sum(value)) as value
            from 
                    inventory_base_report 
            where 
                    date < %s and 
                    ( location_id  IN %s or 
                    warehouse_id IN %s ) and
                     product_id =%s  

            group by product_id


            '''

        


        # caliculating the closing qty.
        closing_qty_query = '''
            select 
                    product_id ,
                    ABS(sum(product_qty)) as qty,
                    ABS(sum(value)) as value
            from 
                    inventory_base_report 
            where 
                    date <= %s and 
                   ( location_id  IN %s or 
                    warehouse_id IN %s ) and
                    product_id =%s and
                    transaction_types NOT IN  ( 'internal' )

            group by product_id

            '''
       
        # caliculating the closing value.
        closing_value_query = '''
            select 
                    product_id ,
                    ABS(sum(product_qty)) as qty,
                    ABS(sum(value)) as value
            from 
                    inventory_base_report 
            where 
                    date <= %s and 
                   ( location_id  IN %s or 
                    warehouse_id IN %s ) and
                    product_id =%s 

            group by product_id

            '''
       
       
        # general
        general_query = '''
            select 
                    product_id ,
                    ABS(sum(product_qty)) as qty,
                    ABS(sum(value)) as value
            from 
                    inventory_base_report 
            where 
                    date >= %s  and 
                    date <= %s  and 
                    ( location_id  IN %s or warehouse_id IN %s ) and
                    product_id =%s and
                    transaction_types = %s  and
                    company_id = %s

            group by product_id

            '''
        # Prouct Price
        product_price ='''
        select * 

        from 
        product_price_history 
        where   
            datetime < %s and 
            product_id = %s
         order by datetime DESC

        '''

        # Prouct Price
        product_price_max ='''
        select * 

        from 
        product_price_history 
        where  
            datetime < %s and  
            product_id = %s
         order by datetime DESC

        '''


        
        # # # GR
        # # # p_receipt
        # # p_receipt_query = '''
        # #     select 
        # #             product_id ,
        # #             ABS(sum(product_qty)) as qty,
        # #             ABS(sum(value)) as value
        # #     from 
        # #             inventory_base_report 
        # #     where 
        # #             date > %s  and 
        # #             date < %s  and 
        # #             ( location_id  IN %s or warehouse_id IN %s ) and
        # #             product_id =%s and
        # #             transaction_types = 'p_receipt' 

        # #     group by product_id

        # #     '''


        # # # p_return
        # # p_return_query = '''
        # #     select 
        # #             product_id ,
        # #             ABS(sum(product_qty)) as qty,
        # #             ABS(sum(value)) as value
        # #     from 
        # #             inventory_base_report 
        # #     where 
        # #             date > %s  and 
        # #             date < %s  and 
        # #             ( location_id  IN %s or warehouse_id IN %s )and
        # #             product_id =%s and
        # #             transaction_types = 'p_return' 

        # #     group by product_id

        # #     '''


        # # # Consumpution qty
        # # m_in_query = '''
        # #     select 
        # #             product_id ,
        # #             ABS(sum(product_qty)) as qty,
        # #             ABS(sum(value)) as value
        # #     from 
        # #             inventory_base_report 
        # #     where 
        # #             date > %s  and 
        # #             date < %s  and 
        # #             ( location_id  IN %s or warehouse_id IN %s )and
        # #             product_id =%s and
        # #             transaction_types = 'in' 

        # #     group by product_id

        # #     '''

        # # # out qty
        # # m_out_query = '''
        # #     select 
        # #             product_id ,
        # #             ABS(sum(product_qty)) as qty,
        # #             ABS(sum(value)) as value
        # #     from 
        # #             inventory_base_report 
        # #     where 
        # #             date > %s  and 
        # #             date < %s  and 
        # #             ( location_id  IN %s or warehouse_id IN %s ) and
        # #             product_id =%s and
        # #             transaction_types = 'out' 

        # #     group by product_id

        # #     '''

        # # # sales s_shipment  quantity
        # # s_shipment_query = '''
        # #     select 
        # #             product_id ,
        # #             ABS(sum(product_qty)) as qty,
        # #             ABS(sum(value)) as value
        # #     from 
        # #             inventory_base_report 
        # #     where 
        # #             date > %s  and 
        # #             date < %s  and 
        # #             ( location_id  IN %s or warehouse_id IN %s ) and
        # #             product_id =%s and
        # #             transaction_types = 's_shipment' 

        # #     group by product_id

        # #     '''


        # # # s_return return quantity 
        # # s_return_query='''
        # #     select 
        # #             product_id ,
        # #             ABS(sum(product_qty)) as qty,
        # #             ABS(sum(value)) as value
        # #     from 
        # #             inventory_base_report 
        # #     where 
        # #             date > %s  and 
        # #             date < %s  and 
        # #             ( location_id  IN %s or warehouse_id IN %s ) and
        # #             product_id =%s and
        # #             transaction_types = 's_return' 

        # #     group by product_id

        # #     '''

        # # # t_shipment
        # # t_shiipment_query = '''
        # #     select 
        # #             product_id ,
        # #             ABS(sum(product_qty)) as qty,
        # #             ABS(sum(value)) as value
        # #     from 
        # #             inventory_base_report 
        # #     where 
        # #             date > %s  and 
        # #             date < %s  and 
        # #             product_id =%s and
        # #             ( location_id  IN %s or warehouse_id IN %s ) and
        # #             transaction_types = 't_shipment' 

        # #     group by product_id
        # #     '''


        # # # t_receipt
        # # t_shiipment_query = '''
        # #     select 
        # #             product_id ,
        # #             ABS(sum(product_qty)) as qty,
        # #             ABS(sum(value)) as value
        # #     from 
        # #             inventory_base_report 
        # #     where 
        # #             date > %s  and 
        # #             date < %s  and 
        # #             product_id =%s and
        # #             ( location_id  IN %s or warehouse_id IN %s ) and
        # #             transaction_types = 't_receipt' 

        # #     group by product_id
        # #     '''
        # # # positive
        # # positive_query='''
        # # select 
        # #             sm.product_id ,sm.product_qty,\
        # #             CASE WHEN ( ib.value < 0 ) THEN (ib.value*-1) ELSE (ib.value) END as value

        # #     from 
        # #             inventory_base_report
        # #     where   
                    
        # #             date > %s  and 
        # #             date < %s and
        # #             ( location_id  IN %s or warehouse_id IN %s ) and
        # #             product_id = %s and
        # #             transaction_types IN ('positive') 

        # #     '''
        # # # (sm.location_id =14 or sm.location_dest_id =14)
        # # # negative
        # # negative_query='''
        # # select 
        # #             sm.product_id ,sm.product_qty,\
        # #             CASE WHEN ( ib.value < 0 ) THEN (ib.value*-1) ELSE (ib.value) END as value

        # #     from 
        # #             inventory_base_report 
        # #     where   
                    
        # #             date > %s  and 
        # #             date < %s and
        # #             ( location_id  IN %s or warehouse_id IN %s ) and
        # #             product_id = %s and
        # #             transaction_types IN ('negative') 

        #     '''

        
        row_c=4
        sl_num=1

        for each_product in self.get_product():


            # based on the user selection finding the warehouse and location.
            if self.warehouse_id:
                curr_warehouse = tuple(self.warehouse_id.ids,)
                curr_location = (0,)
            else:
                curr_warehouse = (0,)
                curr_location = tuple(self.location_id.ids,)


            # opening_qty_query
            opening_params = (self.date_start,curr_location,curr_warehouse,each_product.id)
            self.env.cr.execute(opening_qty_query,opening_params)
            opening_result_qty = self.env.cr.dictfetchall()

            # opening_value_query

            self.env.cr.execute(opening_value_query,opening_params)
            opening_result_vaue = self.env.cr.dictfetchall()

            # product_cost

            # product_params = (self.date_start,each_product.id)
            # self.env.cr.execute(product_price,product_params)
            # op_product_price_result = self.env.cr.dictfetchall()

            # product_cost max
            # product_max_params = (self.date_end,each_product.id)
            # self.env.cr.execute(product_price_max,product_max_params)
            # clo_product_max_price_result = self.env.cr.dictfetchall()


            # if op_product_price_result:
            #     op_product_price_unit = op_product_price_result[0]['cost']
            # else:
            #     op_product_price_unit =0.0

            # if clo_product_max_price_result:
            #     clo_product_price_unit =clo_product_max_price_result[0]['cost']
            # else:
            #     clo_product_price_unit = 0.0



            # closing_qty_query
            closing_params = (self.date_end,curr_location,curr_warehouse,each_product.id)
            self.env.cr.execute(closing_qty_query,closing_params)
            closing_result_qty = self.env.cr.dictfetchall()

            # closing_value_query

            self.env.cr.execute(closing_value_query,closing_params)
            closing_result_value = self.env.cr.dictfetchall()

            # p_receipt
            p_receipt_params = (self.date_start,self.date_end,curr_location,curr_warehouse,each_product.id,'p_receipt',self.company_id.id)
            self.env.cr.execute(general_query,p_receipt_params)
            p_receipt_result = self.env.cr.dictfetchall()

            # p_return
            p_return_params = (self.date_start,self.date_end,curr_location,curr_warehouse,each_product.id,'p_return',self.company_id.id)
            self.env.cr.execute(general_query,p_return_params)
            p_return_result = self.env.cr.dictfetchall()

            # s_shipment
            s_shipment_params = (self.date_start,self.date_end,curr_location,curr_warehouse,each_product.id,'s_shipment',self.company_id.id)
            self.env.cr.execute(general_query,s_shipment_params)
            s_shipment_result = self.env.cr.dictfetchall()

            # s_return
            s_return_params = (self.date_start,self.date_end,curr_location,curr_warehouse,each_product.id,'s_return',self.company_id.id)
            self.env.cr.execute(general_query,s_return_params)
            s_return_result = self.env.cr.dictfetchall()

            # in_query
            m_in_params = (self.date_start,self.date_end,curr_location,curr_warehouse,each_product.id,'in',self.company_id.id)
            self.env.cr.execute(general_query,m_in_params)
            m_in_result = self.env.cr.dictfetchall()

            # out_query
            m_out_params = (self.date_start,self.date_end,curr_location,curr_warehouse,each_product.id,'out',self.company_id.id)
            self.env.cr.execute(general_query,m_out_params)
            m_out_result = self.env.cr.dictfetchall()

            # t_shipment Intra Transfer Shipment
            t_shipment_params = (self.date_start,self.date_end,curr_location,curr_warehouse,each_product.id,'t_shipment',self.company_id.id)
            self.env.cr.execute(general_query,t_shipment_params)
            t_shipment_result = self.env.cr.dictfetchall()

            # # inter_shipment Inter Transfer Shipment
            # inter_shipment_params = (self.date_start,self.date_end,curr_location,curr_warehouse,each_product.id,'inter_shipment',self.company_id.id)
            # self.env.cr.execute(general_query,inter_shipment_params)
            # inter_shipment_params_result = self.env.cr.dictfetchall()

            # t_receipt  Intra Transfer Receipt
            t_receipt_params = (self.date_start,self.date_end,curr_location,curr_warehouse,each_product.id,'t_receipt',self.company_id.id)
            self.env.cr.execute(general_query,t_receipt_params)
            t_receipt_result = self.env.cr.dictfetchall()

            # # inter_receipt  Inter Transfer Receipt
            # inter_receipt_params = (self.date_start,self.date_end,curr_location,curr_warehouse,each_product.id,'inter_receipt',self.company_id.id)
            # self.env.cr.execute(general_query,inter_receipt_params)
            # inter_receipt_result = self.env.cr.dictfetchall()

            # positive
            positive_params = (self.date_start,self.date_end,curr_location,curr_warehouse,each_product.id,'positive',self.company_id.id)
            self.env.cr.execute(general_query,positive_params)
            positive_result = self.env.cr.dictfetchall()

            # negative
            negative_params = (self.date_start,self.date_end,curr_location,curr_warehouse,each_product.id,'negative',self.company_id.id)
            self.env.cr.execute(general_query,negative_params)
            negative_result = self.env.cr.dictfetchall()

            tot_p_qty =re_p_qty = tot_s_qty = tot_m_in_qty =tot_m_out_qty= tot_ts_qty = inter_tot= tot_tr_qty=tot_positive_qty = tot_negative_qty =0.0
            tot_p_value =re_p_value= tot_s_value =tot_m_in_value =tot_m_out_value=tot_ts_value =tot_tr_value=tot_positive_value =tot_negative_value =0.0

            
            # Purchase  = Purchase Receipt 
            if p_receipt_result :
                tot_p_qty = p_receipt_result[0]['qty']  if p_receipt_result else 0.0
                tot_p_value = p_receipt_result[0]['value'] if p_receipt_result else 0.0

            # Purchase  =  Purchase Return
            if p_receipt_result :
                re_p_qty = p_return_result[0]['qty'] if p_return_result else 0.0
                re_p_value =p_return_result[0]['value'] if p_return_result else 0.0

            # Sales = Sales Shipment - Sales Return
            if s_shipment_result :
                tot_s_qty = s_shipment_result[0]['qty'] 
                tot_s_value = s_shipment_result[0]['value']
           
            #  Manufacture (in)
            if m_in_result :
                tot_m_in_qty = (m_in_result[0]['qty'] if m_in_result else 0.0)
                tot_m_in_value =(m_in_result[0]['value'] if m_in_result else 0.0)

            #  Consumed (out)
            if m_out_result :
                tot_m_out_qty = (m_out_result[0]['qty'] if m_out_result else 0.0)
                tot_m_out_value =(m_out_result[0]['value'] if m_out_result else 0.0)
           
            # Consumed = Transfer Shipment- Transfer Receipt
            if t_shipment_result :
                tot_ts_qty =(t_shipment_result[0]['qty'] if t_shipment_result else 0.0)
                tot_ts_value = (t_shipment_result[0]['value'] if t_shipment_result else 0.0)
            
            if t_receipt_result :
                tot_tr_qty =(t_receipt_result[0]['qty'] if t_receipt_result else 0.0)
                tot_tr_value =  (t_receipt_result[0]['value'] if t_receipt_result else 0.0)
            
            # Positive
            if positive_result:
                tot_positive_qty=positive_result[0]['qty'] if positive_result else 0.0
                tot_positive_value=positive_result[0]['value'] if positive_result else 0.0

            # Positive
            if negative_result:
                tot_negative_qty=negative_result[0]['qty'] if negative_result else 0.0
                tot_negative_value=negative_result[0]['value'] if negative_result else 0.0


            sl_val = ws.cell(row=row_c, column=1, value=sl_num)
            sl_val.alignment = Alignment(horizontal='center',vertical='center')


            product_id = ws.cell(row=row_c, column=2, value=each_product.name)


            product_cat = ws.cell(row=row_c, column=3, value=each_product.categ_id.name)

            product_code = ws.cell(row=row_c, column=4, value=each_product.default_code)

            # location_name = ws.cell(row=row_c, column=5, value=product[0]['uu_name'])

            p_unit = ws.cell(row=row_c, column=5, value=each_product.uom_id.name)

            sys_opening = ws.cell(row=row_c, column=6, value= opening_result_qty[0]['qty'] if opening_result_qty else 0.0 )
            sys_amt = ws.cell(row=row_c, column=7, value= (opening_result_vaue[0]['value'] if opening_result_vaue else 0.0)  )

            purchase_qty = ws.cell(row=row_c, column=8, value=tot_p_qty)
            purchase_amt = ws.cell(row=row_c, column=9, value=tot_p_value)

            purchase_r_qty = ws.cell(row=row_c, column=10, value=re_p_qty)
            purchase_r_amt = ws.cell(row=row_c, column=11, value=re_p_value)

            sale_qty = ws.cell(row=row_c, column=12, value=tot_s_qty )
            sale_amt = ws.cell(row=row_c, column=13, value=tot_s_value )

            sale_re_qty = ws.cell(row=row_c, column=14, value=s_return_result[0]['qty'] if s_return_result else 0.0 )
            sale_re_amt = ws.cell(row=row_c, column=15, value=s_return_result[0]['value'] if s_return_result else 0.0 )
            
            mrp_qty = ws.cell(row=row_c, column=16, value=tot_m_out_qty )
            mrp_amt = ws.cell(row=row_c, column=17, value=tot_m_out_value )

            consumed_qty = ws.cell(row=row_c, column=18, value=tot_m_in_qty )
            consumed_amt = ws.cell(row=row_c, column=19, value=tot_m_in_value )

            postive_qty = ws.cell(row=row_c, column=20, value=tot_positive_qty )
            postive_amt = ws.cell(row=row_c, column=21, value=tot_positive_value )

            negative_qty = ws.cell(row=row_c, column=22, value=tot_negative_qty )
            negative_amt = ws.cell(row=row_c, column=23, value=tot_negative_value )



            t_shipment_inter_qty = ws.cell(row=row_c, column=24, value= tot_ts_qty)
            t_shipment_inter_amt = ws.cell(row=row_c, column=25, value= tot_ts_value)

            # t_shipment_intra_qty = ws.cell(row=row_c, column=24, value=tot_ts_qty )
            # t_shipment_intra_amt = ws.cell(row=row_c, column=25, value=tot_ts_value )

            t_receipt_inter_qty = ws.cell(row=row_c, column=26, value= tot_tr_qty)
            t_receipt_inter_amt = ws.cell(row=row_c, column=27, value= tot_tr_value)

            # t_receipt_intra_qty = ws.cell(row=row_c, column=28, value=tot_tr_qty )
            # t_receipt_intra_amt = ws.cell(row=row_c, column=29, value=tot_tr_value )

            closing_qty = ws.cell(row=row_c, column=28, value=closing_result_qty[0]['qty'] if closing_result_qty else 0.0 )
            closing_amt = ws.cell(row=row_c, column=29, value=(closing_result_value[0]['value']if closing_result_value else 0.0) )

            sl_num +=1
            row_c +=1



        fp = io.BytesIO()
        wb.save(fp)
        excel_file = base64.encodestring(fp.getvalue())
        self.inv_report = excel_file
        self.inv_printed = True
        self.file_name = "Inventory Report.xlsx"
        fp.close()

        return {
        'view_mode': 'form',
        'res_id': self.id,
        'res_model': 'inventory.report.excel',
        'view_type': 'form',
        'type': 'ir.actions.act_window',
        'context': self.env.context,
        'target': 'new',
                   }


   
